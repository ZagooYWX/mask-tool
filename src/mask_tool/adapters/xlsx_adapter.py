"""Excel文档(.xlsx)适配器"""

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from mask_tool.adapters.base import FileAdapter
from mask_tool.models.detection import DetectionResult, DetectionStatus, Location


class XlsxAdapter(FileAdapter):
    """Excel文档脱敏适配器"""

    def supported_extensions(self) -> list[str]:
        return [".xlsx"]

    def process(self, input_path: Path, output_dir: Path) -> Path:
        """
        处理Excel文档：
        - 逐工作表、逐单元格处理
        - 保持格式不变
        """
        wb = load_workbook(str(input_path))
        file_name = input_path.stem
        output_path = output_dir / f"{file_name}_masked.xlsx"

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value and isinstance(cell.value, str):
                        self._process_cell(cell, str(input_path), sheet_name)

        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    def _process_cell(self, cell, file_path: str, sheet_name: str) -> None:
        """处理单个单元格"""
        text = str(cell.value)
        if not text.strip():
            return

        cell_ref = cell.coordinate  # 如 "A3", "B5"

        # 检测
        results = self.detector.detect(text, file_path)

        # 设置位置信息
        for r in results:
            r.location.sheet = sheet_name
            r.location.cell_ref = cell_ref

        # 策略决策
        results = self.policy.apply(results)

        # 执行替换
        masked_text = text
        for result in results:
            if result.status not in (DetectionStatus.AUTO_MASK, DetectionStatus.SUGGEST_MASK):
                continue
            if result.text not in masked_text:
                continue

            if self.masker.irreversible:
                replacement = "***"
            else:
                from mask_tool.models.mapping import TokenMapping
                replacement = self.masker.token_gen.generate(result.text, result.text_type)
                self.masker.mappings.append(TokenMapping(
                    token=replacement,
                    original=result.text,
                    text_type=result.text_type,
                    confidence=result.confidence,
                ))

            masked_text = masked_text.replace(result.text, replacement)

        if masked_text != text:
            cell.value = masked_text
